import pytest
from openpyxl import load_workbook

EXCEL_FILE = './tests/data/excel_test_data.xlsx'

# pytest discover tests if function names have "test" in the beginning.
# c.f. https://docs.pytest.org/en/latest/explanation/goodpractices.html#test-discovery
@pytest.fixture(scope='module')
def report():
    workbook = load_workbook(filename=EXCEL_FILE, read_only=True, data_only=True)
    worksheet = workbook['Sheet1']
    return worksheet


def test_read_a1_excel(report):
    assert report.cell(row=1, column=1).value == 'Cell A1!'
